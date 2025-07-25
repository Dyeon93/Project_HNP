from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import pandas as pd
import urllib.parse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time

search_word = input("검색어 입력: ")
encoded = urllib.parse.quote(search_word)

url = f"https://www.coupang.com/np/search?component=&q={encoded}&channel=user"

max_items = 100
scroll_pause = 1.5

collected_info = set()
collected_data = []

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto(url)
    page.wait_for_load_state("networkidle")
    time.sleep(5)

    while len(collected_data) < max_items:
        html = page.content()
        soup = BeautifulSoup(html, "html.parser")

        # 상품 항목 추출
        items = soup.find_all("li", class_="ProductUnit_productUnit__Qd6sv")
        for item in items:
            # ✅ 광고 상품 제외
            if item.find(class_="AdMark_text__Rp7px"):
                continue
            # 제품 링크
            link_tag = item.find('a', href=True)
            if link_tag is None:
                continue
            product_link = link_tag['href']

            # 이미지 URL
            img_tag = item.find("img", src=True)
            if img_tag is None:
                continue
            image_url = img_tag['src']

            if product_link not in collected_info:
                collected_info.add(product_link)

                name_tag = item.find(class_="ProductUnit_productName__gre7e")
                basePrice_tag = item.find(class_="PriceInfo_basePrice__8BQ32")
                discount_tag = item.find(class_="PriceInfo_discountRate__EsQ8I")
                price_tag = item.find(class_="Price_priceValue__A4KOr")
                rating_tag = item.find(class_="ProductRating_star__RGSlV")
                review_tag = item.find(class_="ProductRating_ratingCount__R0Vhz")

                collected_data.append({
                    'image': image_url,
                    'name': name_tag.text.strip() if name_tag else '-',
                    'basePrice': basePrice_tag.text.strip() if basePrice_tag else '-',
                    'discount': discount_tag.text.strip() if discount_tag else '0%',
                    'price': price_tag.text.strip() if price_tag else '-',
                    'rating': rating_tag.text.strip() if rating_tag else '-',
                    'review_count': review_tag.text.strip().replace('(', '').replace(')', '') if review_tag else '0',
                    'link': "https://www.coupang.com" + product_link
                })

                print(f"현재 수집된 상품 수: {len(collected_data)}")

            if len(collected_data) >= max_items:
                print("✅ 100개 수집 완료!")
                break

        # 다음 페이지로 이동 (data-page 기반 자동 탐색)
        try:
            current_page_tag = page.query_selector('a[class*="Pagination_selected"]')
            if current_page_tag is None:
                print("❌ 현재 페이지 태그를 찾을 수 없습니다.")
                break

            current_page_num = int(current_page_tag.get_attribute('data-page'))
            next_page_num = current_page_num + 1

            next_page_tag = page.query_selector(f'a[data-page="{next_page_num}"]')
            if next_page_tag is None:
                print("❌ 다음 페이지 버튼을 찾을 수 없습니다. 종료합니다.")
                break

            next_page_tag.click()
            time.sleep(2)
        except Exception as e:
            print(f"❌ 다음 페이지 이동 실패: {e}")
            break

    browser.close()

df = pd.DataFrame(collected_data)

# === 1. 크롤링 순위 추가 및 컬럼 순서 조정 ===
df["rank"] = df.index + 1
cols = df.columns.tolist()
cols.insert(0, cols.pop(cols.index("rank")))
df = df[cols]

# === 2. 파일 저장 ===
lt = time.localtime()
filename = f"{lt.tm_year}.{lt.tm_mon}.{lt.tm_mday}_쿠팡_{search_word}_상품정보_Playwright.xlsx"
df.to_excel(filename, index=False)

# === 3. openpyxl 로드 ===
wb = load_workbook(filename)
ws = wb.active

# === 4. C열(name)에 하이퍼링크 삽입 ===
for row in range(2, 2 + len(df)):
    name = df.loc[row - 2, "name"]
    link = df.loc[row - 2, "link"]

    cell = ws.cell(row=row, column=3)  # C열
    cell.value = name
    cell.hyperlink = link
    cell.style = "Hyperlink"

# === 5. 'link' 컬럼 제거 ===
link_col_index = df.columns.get_loc("link") + 1
ws.delete_cols(link_col_index)

# === 6. 셀 너비 자동 조절 + B열은 고정 ===
for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
    col_letter = get_column_letter(col_idx)

    if col_letter == "B":
        ws.column_dimensions[col_letter].width = 13  # B열은 고정
        continue

    max_length = 0
    for cell in column_cells:
        try:
            cell_value = str(cell.value)
            max_length = max(max_length, len(cell_value))
        except:
            pass

    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

# === 7. 행 높이 조정 (1행 제외) ===
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 70  # 데이터 행만 높이 조정

# === 8. 셀 정렬 설정 (가로 왼쪽, 세로 가운데) ===
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(vertical='center')

# === 9. 저장 완료 ===
wb.save(filename)

print(f"\n✅ 엑셀 저장 완료! → {filename}")