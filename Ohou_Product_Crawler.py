from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import pandas as pd
import urllib.parse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import time

search_word = input("검색어 입력: ")
encoded = urllib.parse.quote(search_word)

url = f"https://ohou.se/productions/feed?query={encoded}&search_affect_type=Typing"

max_items = 100
scroll_pause = 0.8
max_same_height = 3  # 동일 scrollHeight가 몇 번 반복되면 종료할지

collected_info = set()
collected_links = set()
collected_data = []

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    page.goto(url)
    time.sleep(2)

    prev_height = 0
    same_height_count = 0

    # 스크롤 반복
    while len(collected_data) < max_items:
        curr_height = page.evaluate("() => document.body.scrollHeight")

        if curr_height == prev_height:
            same_height_count += 1
            print(f"⚠️ 스크롤 높이 동일 ({same_height_count}회)")
            if same_height_count >= max_same_height:
                print("📌 더 이상 페이지 높이 변화 없음. 종료합니다.")
                break
        else:
            same_height_count = 0

        prev_height = curr_height

        # 스크롤
        page.keyboard.press("PageDown")
        time.sleep(scroll_pause)

        # 파싱
        html = page.content()
        soup = BeautifulSoup(html, "html.parser")
        items = soup.find_all("div", class_="production-feed__item-wrap col-6 col-md-4 col-lg-3")

        for item in items:
            link_tag = item.find('a', href=True)
            if link_tag is None:
                continue
            product_link = link_tag['href']

            img_tag = item.find(class_='thumbnail-image e1bro5mc1 css-7bfh27', src=True)
            if img_tag is None:
                continue
            image_url = img_tag.get('src')

            if product_link not in collected_info:
                collected_info.add(product_link)

                brand_name = item.find(class_=['product-brand', 'css-11vbb10', 'eo39oc45'])
                product_name = item.find(class_=['product-name', 'css-11e7usa', 'eo39oc44'])
                discount = item.find(class_=['css-nqvy3g', 'e175igv96'])
                price = item.find(class_=['css-13aof0h', 'e175igv95'])
                rating = item.find(class_='avg')
                review_count = item.find(class_='count')

                collected_data.append({
                    'image': image_url,
                    'name': product_name.text.strip() if product_name else '-',
                    'brand': brand_name.text.strip() if brand_name else '-',
                    'discount': discount.text.strip() if discount else '0%',
                    'price': price.text.strip() if price else '-',
                    'rating': rating.text.strip() if rating else '-',
                    'review_cnt': review_count.text.strip().replace('(', '').replace(')', '') if review_count else '0',
                    'link': 'https://ohou.se' + product_link
                })

                print(f"현재 수집된 상품 수: {len(collected_data)}")

            if len(collected_data) >= max_items:
                print("✅ 100개 수집 완료!")
                break

    browser.close()

# === DataFrame 및 엑셀 정리 ===
df = pd.DataFrame(collected_data)

# 1. 순위 컬럼 추가
df["rank"] = df.index + 1
cols = df.columns.tolist()
cols.insert(0, cols.pop(cols.index("rank")))
df = df[cols]

# 2. 엑셀 저장
lt = time.localtime()
filename = f"{lt.tm_year}.{lt.tm_mon}.{lt.tm_mday}_오늘의집_{search_word}_상품정보.xlsx"
df.to_excel(filename, index=False)

# 3. 엑셀 하이퍼링크 및 서식 처리
wb = load_workbook(filename)
ws = wb.active

# 4. C열(name)에 하이퍼링크 삽입
for row in range(2, 2 + len(df)):
    name = df.loc[row - 2, "name"]
    link = df.loc[row - 2, "link"]

    cell = ws.cell(row=row, column=3)  # C열
    cell.value = name
    cell.hyperlink = link
    cell.style = "Hyperlink"

# 5. 'link' 컬럼 제거
link_col_index = df.columns.get_loc("link") + 1
ws.delete_cols(link_col_index)

# 6. 셀 너비 조정
for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
    col_letter = get_column_letter(col_idx)

    if col_letter == "B":
        ws.column_dimensions[col_letter].width = 13
        continue

    max_length = 0
    for cell in column_cells:
        try:
            cell_value = str(cell.value)
            max_length = max(max_length, len(cell_value))
        except:
            pass

    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

# 7. 행 높이 조정
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 70

# 8. 정렬
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(vertical='center')

# 9. 저장
wb.save(filename)
print(f"\n✅ 엑셀 저장 완료! → {filename}")
